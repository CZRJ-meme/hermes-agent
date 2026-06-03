#!/usr/bin/env python3
import openpyxl, re, json

wb = openpyxl.load_workbook('/Users/czrj/Downloads/升级小游戏_埋点方案_20260523.xlsx', read_only=True)
ws = wb['#事件数据']
new_events = {}
current = None
for row in ws.iter_rows(values_only=True):
    if row[0]:
        current = row[0]
        new_events[current] = {'display': row[1], 'desc': row[2], 'module': row[3], 'props': []}
    if row[4] and current:
        new_events[current]['props'].append(row[4])

with open('/Users/czrj/.hermes/skills/thinkingdata-te/references/event-tracking-schema.md') as f:
    old = f.read()

old_events = set(re.findall(r'`([A-Z][A-Za-z0-9_]+)`', old))

key_events = [
    'ConnectionServiceLogin', 'ConnectionServiceLogout', 'RoundMatchRequestAccept',
    'BackPurchaseSdkOK', 'BackPurchaseSdkFail', 'BackPurchaseSdkCreate', 'CreateRechargeOrderFail',
    'BackPurchaseRegisterSuccess', 'BackPurchaseLoginSuccess', 'BackPurchaseConsumables', 'BackPurchaseGoodsSend',
    'RoundStart', 'RoundEnd', 'RoundInitialCard', 'RoundGrabMainCard', 'RoundDoubleReward',
    'AdVideoStart', 'AdVideoComplete', 'AdVideoFail',
    'MallPurchaseSuccess', 'MallPurchaseFail',
    'GiftpkgSendLog', 'RechargeTriggerLimit', 'RoundGrabMainCardEnd', 'RoudTurnHoleCard',
    'UserAbnormalLogin', 'GameUserAbnormalLogin', 'kms_complete_recharge',
    'RoundOver', 'RoundMatchRequestAccept',
]

print('Key events comparison:')
for evt in key_events:
    in_new = evt in new_events
    in_old = evt in old_events
    status = 'OK' if in_new else 'MISSING'
    yn = 'y' if in_old else 'n'
    print(f'  {evt}: {status} (in_old={yn})')
    if in_new:
        d = new_events[evt]
        print(f'    -> {d["display"]} [{d["module"]}]')
        if d['props']:
            print(f'    -> {len(d["props"])} props: {d["props"][:6]}')

# Client events
py_events = [k for k in new_events.keys() if k.startswith('c_')]
print(f'\nClient events (c_*): {len(py_events)}')
for e in sorted(py_events)[:15]:
    print(f'  {e}: {new_events[e]["display"]}')

# Compare user properties
ws_user = wb['#用户数据']
new_user_props = {}
for row in ws_user.iter_rows(values_only=True):
    if row[0] and row[0] != '属性名（必填）':
        new_user_props[row[0]] = {'display': row[1], 'type': row[2], 'update': row[3], 'desc': row[4]}

print(f'\nUser properties: {len(new_user_props)}')
for name, data in sorted(new_user_props.items()):
    print(f'  {name}: {data["display"]} ({data["type"]}, {data["update"]})')

# Compare public event properties
ws_pub = wb['#公共事件属性']
new_pub_props = {}
for row in ws_pub.iter_rows(values_only=True):
    if row[0] and row[0] != '属性名（必填）':
        new_pub_props[row[0]] = {'display': row[1], 'type': row[2], 'desc': row[3]}

print(f'\nPublic event properties: {len(new_pub_props)}')
# Check key ones
for key in ['channel_name', 'app_id', 'app_name', 'app_version', '#os', '#os_version', '#device_id', '#device_model', '#city', '#province', '#country', '#network_type', '#carrier']:
    if key in new_pub_props:
        d = new_pub_props[key]
        print(f'  {key}: {d["display"]} ({d["type"]})')
    else:
        print(f'  {key}: NOT FOUND')
